from openpyxl import load_workbook
import requests

from config_settings import ENV, IMPORT_FN


class vehicle(object):
    def __init__(self, row):
        self.vehicle_id   = str(row[0])
        self.name     = str(row[1])
        self.capacity = int(row[2])
        self.speed    = float(row[3])

    def __repr__(self):
        return '{0} [ID: {1}]'.format(self.name, self.vehicle_id)

    def update_vf(self, ENV, COOKIE):
        tmp = requests.put('https://{0}.versafleet.co/vehicles/'.format(ENV) + self.vehicle_id,
                           headers = {
                               'cookie': COOKIE
                           },
                           json = {
                               'vehicle': {
                                   'plate_number':self.name,
                                   'cargo_load': self.capacity,
                                   'speed': self.speed
                               }
                           }
        )
        if tmp.status_code == 200:
            print self, tmp
            return None
        else:
            print "\nERROR WITH {0}\n{1}\n{2}\n".format(self, tmp, tmp.text)
            return tmp


class driver(object):
    def __init__(self, row):
        TIMESLOTS = {
            '0': 780,
            '1': 4380,
            '2': 7980,
            '4': 15180,
        }
        self.driver_id  = str(row[0])
        self.name       = str(row[1])
        self.paired_veh = str(row[2])
        self.vd_status  = row[3]
        self.username   = str(row[4])
        self.password   = str(row[5])
        self.wh_start   = TIMESLOTS[str(row[6])]

    def __repr__(self):
        return '{0} [ID: {1}]'.format(self.name, self.driver_id)

    def update_vf(self, ENV, COOKIE, vehicle_dict):
        JSON = {'driver': {
            'name':self.name,
            'default_vehicle_id':vehicle_dict[self.paired_veh].vehicle_id,
            'is_versadrive_user':self.vd_status,
            'username':""
        }}
        if self.vd_status:
            JSON['driver']['username'] = self.username
            JSON['driver']['password'] = self.password

        tmp = requests.put('https://{0}.versafleet.co/drivers/'.format(ENV) + self.driver_id,
                           headers = {
                               'cookie': COOKIE
                           },
                           json = JSON
        )
        if tmp.status_code == 200:
            print self, tmp
            return None
        else:
            print "\nERROR WITH {0}\n{1}\n{2}\n".format(self, tmp, tmp.text)
            return tmp

    def update_wh_start(self, ENV, COOKIE):
        tmp_in = requests.get('https://{0}.versafleet.co/drivers/'.format(ENV) + self.driver_id,
                              headers = {
                                  'cookie': COOKIE
                              })
        WH_JSON = tmp_in.json()['driver']['working_hours']
        NEW_WH_JSON = {'driver':{'working_hours_attributes':[]}}
        for wh in WH_JSON:
            wh['start_time'] = self.wh_start
            NEW_WH_JSON['driver']['working_hours_attributes'].append(wh)

        tmp = requests.put('https://{0}.versafleet.co/drivers/'.format(ENV) + self.driver_id,
                           headers = {
                               'cookie': COOKIE
                           },
                           json = NEW_WH_JSON
        )

        if tmp.status_code == 200:
            print self, tmp
            return None
        else:
            print "\nERROR WITH {0}\n{1}\n{2}\n".format(self, tmp, tmp.text)
            return tmp

def read_excel(file_n):
    '''
    Takes an excel sheet with 2 sheeets,
    VEHICLE
    Veh ID | Vehicle Name | Capacity | Speed

    DRIVER
    Driver ID | Driver Name | Vehicle Name | VersaDrive User | Username | Password | Time Slot
    '''
    driver_list = []
    vehicle_dict = {}
    wb = load_workbook(file_n, read_only=True)

    vehicle_sheet = wb.worksheets[0]
    for row in vehicle_sheet.iter_rows(min_row=2):
        vehicle_dict[str(row[1].value)] = vehicle([cell.value for cell in row])

    driver_sheet = wb.worksheets[1]
    for row in driver_sheet.iter_rows(min_row=2):
        driver_list.append(driver([cell.value for cell in row]))

    return driver_list, vehicle_dict




# Get Cookie
COOKIE = raw_input("Login to VF to get a cookie and input it here:\n>>").strip()

# Open excel file, get vehicle & driver actions
driver_list, vehicle_dict = read_excel(IMPORT_FN)

# perform vehicle actions
vehicle_error_dict = {}

if raw_input("\nKey 'N' to skip vehicle update\n>>").upper().strip() != 'N':
    for k, veh in vehicle_dict.iteritems():
        tmp = veh.update_vf(ENV, COOKIE)
        if tmp != None:
            vehicle_error_dict[k] = veh


# perform driver actions
driver_error_list = []

if raw_input("\nKey 'N' to skip driver update\n>>").upper().strip() != 'N':
    for driver in driver_list:
        tmp = driver.update_vf(ENV, COOKIE, vehicle_dict)
        if tmp != None:
            driver_error_list.append(driver)


driver_wh_error_list = []

if raw_input("\nKey 'N' to skip driver working hour update\n>>").upper().strip() != 'N':
    for driver in driver_list:
        tmp = driver.update_wh_start(ENV, COOKIE)
        if tmp != None:
            driver_error_list.append(driver)
